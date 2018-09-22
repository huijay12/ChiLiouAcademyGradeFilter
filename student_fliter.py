import openpyxl
from operator import itemgetter, attrgetter

refer_file = 'result_TW2017.xlsx'
output_file = '重修生名單.xlsx'

name_sheet_grade1 = '2017-修養科'
name_sheet_grade2 = '2017-本科'
name_sheet_grade3 = '2017-研究科'

out_name_sheet_grade1 = '2017~~修重'
out_name_sheet_grade2 = '2017~~本重 '
out_name_sheet_grade3 = '2017~~研究重 '

re_start_cell = 'A2'
re_end_cell = 'R257'  #修257  本183  研159

out_start_row = 2
out_row_block = 21

out_existed_end_row = 126  #修126  本63  研71
out_last_num = 239   #修239  本120  研127

stu_list = []


class Stu:
    def __init__(self, sn, n, a, e, p, c):
        self.stunum = sn
        self.name = n
        self.attn = a
        self.exam = e
        self.prac = p
        self.comment = c



STU_NUM = 0
NAME = 2
COMMEMT = 17
PASS_ATTN = 9
PASS_EXAM = 12
PASS_PRACT = 15


refer_wb = openpyxl.load_workbook(refer_file, data_only = True)    #載入excel檔
output_wb = openpyxl.load_workbook(output_file)

sheet1 = refer_wb.get_sheet_by_name(name_sheet_grade1)  #抓到某個分頁
out_sheet1 = output_wb.get_sheet_by_name(out_name_sheet_grade1)

refer_data = tuple(sheet1[re_start_cell:re_end_cell])        #抓到特定矩形區塊資料, 存成tuple


for r_data in refer_data:

    stunum = r_data[STU_NUM].value
    stuname = r_data[NAME].value

    if r_data[COMMEMT].value == 'O.K.':   #看重修名單裡有沒有名字  有的話刪掉
    
        for i in range(out_start_row, out_existed_end_row+1):

            if out_sheet1.cell(row=i, column=2).value == stunum:   #左行
                out_sheet1.cell(row=i, column=2).value = ' '
                out_sheet1.cell(row=i, column=2+1).value = ' '
                out_sheet1.cell(row=i, column=2+2).value = ' '
                out_sheet1.cell(row=i, column=2+3).value = ' '
                out_sheet1.cell(row=i, column=2+4).value = ' '

            if out_sheet1.cell(row=i, column=10).value == stunum:   #右行
                out_sheet1.cell(row=i, column=10).value = ' '
                out_sheet1.cell(row=i, column=10+1).value = ' '
                out_sheet1.cell(row=i, column=10+2).value = ' '
                out_sheet1.cell(row=i, column=10+3).value = ' '
                out_sheet1.cell(row=i, column=10+4).value = ' '


    
    if r_data[COMMEMT].value != 'O.K.':
#        print("PASS_ATTN: " + r_data[PASS_ATTN].value + "  PASS_EXAM: " + r_data[PASS_EXAM].value + "  PASS_PRACT: " + r_data[PASS_PRACT].value)
    
        existed = 0;

        #先找有沒有已存在於名單裡
        for i in range(out_start_row, out_existed_end_row+1):

            if out_sheet1.cell(row=i, column=2).value == stunum:   #左行

                if r_data[PASS_ATTN].value == 'No':
                    out_sheet1.cell(row=i, column=2+2).value = '0'
                else:
                    out_sheet1.cell(row=i, column=2+2).value = ' '

                if r_data[PASS_EXAM].value == 'No':
                    out_sheet1.cell(row=i, column=2+3).value = '0'
                else:
                    out_sheet1.cell(row=i, column=2+3).value = ' '

                if r_data[PASS_PRACT].value == 'No':
                    out_sheet1.cell(row=i, column=2+4).value = '0'
                else:
                    out_sheet1.cell(row=i, column=2+4).value = ' '

                existed = 1;


            if out_sheet1.cell(row=i, column=10).value == stunum:   #右行

                if r_data[PASS_ATTN].value == 'No':
                    out_sheet1.cell(row=i, column=10+2).value = '0'
                else:
                    out_sheet1.cell(row=i, column=10+2).value = ' '

                if r_data[PASS_EXAM].value == 'No':
                    out_sheet1.cell(row=i, column=10+3).value = '0'
                else:
                    out_sheet1.cell(row=i, column=10+3).value = ' '

                if r_data[PASS_PRACT].value == 'No':
                    out_sheet1.cell(row=i, column=10+4).value = '0'
                else:
                    out_sheet1.cell(row=i, column=10+4).value = ' '

                existed = 1;


        if existed == 0:

            out_last_num = out_last_num+1
            
            if out_last_num%20 == 0:
                if out_last_num%40 == 0:
                    row = out_row_block*(int(out_last_num/40))
                else:
                    row = out_row_block*(int(out_last_num/40)) + 1 + 20
            else:
                row = out_row_block*(int(out_last_num/40)) + 1 + (out_last_num%20)

            if out_last_num % 40 < 21:
                col = 1
            else:
                col = 9

            if out_last_num % 40 == 0:
                col = 9


            out_sheet1.cell(row=row, column=col).value = out_last_num
            out_sheet1.cell(row=row, column=col+1).value = stunum
            out_sheet1.cell(row=row, column=col+2).value = stuname

            if r_data[PASS_ATTN].value == 'No':
                out_sheet1.cell(row=row, column=col+3).value = '0'

            if r_data[PASS_EXAM].value == 'No':
                out_sheet1.cell(row=row, column=col+4).value = '0'

            if r_data[PASS_PRACT].value == 'No':
                out_sheet1.cell(row=row, column=col+5).value = '0'




if out_last_num%20 == 0:
    if out_last_num%40 == 0:
        out_existed_end_row = out_row_block*(int(out_last_num/40))
    else:
        out_existed_end_row = out_row_block*(int(out_last_num/40)) + 1 + 20

for i in range(1, out_last_num):

    if out_sheet1.cell(row=i, column=2).value != ' ' and out_sheet1.cell(row=i, column=2).value != '院生番号' and out_sheet1.cell(row=i, column=2).value !=  None:
        
        stunum = out_sheet1.cell(row=i, column=2).value
        name = out_sheet1.cell(row=i, column=2+1).value
        attn = out_sheet1.cell(row=i, column=2+2).value
        exam = out_sheet1.cell(row=i, column=2+3).value
        prac = out_sheet1.cell(row=i, column=2+4).value
        comment = out_sheet1.cell(row=i, column=2+5).value
        stu_list.append(Stu(stunum, name, attn, exam, prac, comment))


    if out_sheet1.cell(row=i, column=10).value != ' ' and out_sheet1.cell(row=i, column=10).value != '院生番号' and out_sheet1.cell(row=i, column=10).value !=  None:
        
        stunum = out_sheet1.cell(row=i, column=10).value
        name = out_sheet1.cell(row=i, column=10+1).value
        attn = out_sheet1.cell(row=i, column=10+2).value
        exam = out_sheet1.cell(row=i, column=10+3).value
        prac = out_sheet1.cell(row=i, column=10+4).value
        comment = out_sheet1.cell(row=i, column=10+5).value
        stu_list.append(Stu(stunum, name, attn, exam, prac, comment))


stu_list = sorted(stu_list, key=attrgetter('stunum'))


for i in range(1, len(stu_list)):

    print(str(i) + ' ' + str(stu_list[i].stunum))

#print(str(len(stu_list)))


#for i in range(1, 200):
#    for j in range(0, 7):
#        out_sheet1.cell(row=i, column=1+j).value = None
#        out_sheet1.cell(row=i, column=9+j).value = None


for i in range(1, len(stu_list)):

    if i%20 == 0:
        if i%40 == 0:
            row = out_row_block*(int(i/40))
        else:
            row = out_row_block*(int(i/40)) + 1 + 20
    else:
        row = out_row_block*(int(i/40)) + 1 + (i%20)

    if i % 40 < 21:
        col = 1
    else:
        col = 9

    if i % 40 == 0:
        col = 9

    out_sheet1.cell(row=row, column=col+1).value = stu_list[i].stunum
    out_sheet1.cell(row=row, column=col+2).value = stu_list[i].name
    out_sheet1.cell(row=row, column=col+3).value = stu_list[i].attn
    out_sheet1.cell(row=row, column=col+4).value = stu_list[i].exam
    out_sheet1.cell(row=row, column=col+5).value = stu_list[i].prac
    out_sheet1.cell(row=row, column=col+6).value = stu_list[i].comment


print(str(out_last_num))

for i in range(len(stu_list), out_last_num+1):

    if i%20 == 0:
        if i%40 == 0:
            row = out_row_block*(int(i/40))
        else:
            row = out_row_block*(int(i/40)) + 1 + 20
    else:
        row = out_row_block*(int(i/40)) + 1 + (i%20)

    if i % 40 < 21:
        col = 1
    else:
        col = 9

    if i % 40 == 0:
        col = 9

    for j in range(0, 7):
        out_sheet1.cell(row=row, column=col+j).value = None
        out_sheet1.cell(row=row, column=col+j).value = None




output_wb.save(output_file)